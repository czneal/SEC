# -*- coding: utf-8 -*-
"""
Created on Thu Nov 15 13:12:13 2018

@author: Asus
"""

import pandas as pd
from openpyxl import load_workbook
import numpy as np

#us-gaap:LiabilitiesCurrent	    mg:AccruedLiabilitiesDeferredRevenuesAndContingentConsiderationLiabilitiesCurrent

#adshs = pd.read_csv('outputs/diffliabs/2018-11-13/liab_leaf.csv')
def autoreplace():
    adshs = set()
    with open('outputs/diffliabs/2018-11-13/liab_redeemable.csv') as f:
        adshs.update([l.replace('\n', '') for l in f.readlines()])


    wb = load_workbook('outputs/diffliabs/2018-11-13/liab_lcpc_m_new_sum_c.xlsx')
    ws = wb.active

    for row in range(1,1575):
        if str(ws['A'+str(row)].value) in adshs:
            ws['K' + str(row)] = 'added us-gaap:RedeemableNoncontrollingInterestEquityCarryingAmount'
    wb.save('outputs/diffliabs/2018-11-13/liab_lcpc_m_new_sum_c.xlsx')

def multiplyautoreplace():
    adshs = dict()
    with open('outputs/diffliabs/2018-11-13/liab_wrong_class.csv') as f:
        for l in f.readlines():
            (k, v) = l.replace('\n', '').split('\t')
            if k in adshs:
                adshs[k].append(v.strip())
            else:
                adshs[k] = [v.strip()]

    wb = load_workbook('outputs/diffliabs/2018-11-13/liab_lcpc_m_new_sum_c.xlsx')
    ws = wb.active

    for row in range(1,1575):
        k = str(ws['A'+str(row)].value)
        if k in adshs:
            message = 'missed'
            for v in adshs[k]: message += ' ' + v
            ws['K' + str(row)] = message

    wb.save('outputs/diffliabs/2018-11-13/liab_lcpc_m_new_sum_c.xlsx')